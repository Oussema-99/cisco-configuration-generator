import openpyxl
from jinja2 import Template
from datetime import datetime
import os


# Pull the Excel File and read the Sheets
def excel_pull(excel_path: str):
	workbook = openpyxl.load_workbook(excel_path)
	time = datetime.now().strftime("%m_%d_%Y_%H_%M_%S")
	directory_path = f"output/{time}/"
	os.mkdir(directory_path)
	
	for sheet in workbook.worksheets:
		read_sheet(sheet, directory_path)


# Pull Data from rows and insert it into "row_data" one by one
def read_sheet(sheet, directory: str):
	nbr_rows = sheet.max_row + 1
	nbr_col = sheet.max_column + 1
	for i in range(2, nbr_rows):
		jinja2_template = sheet.cell(row=i, column=1).value
		if not jinja2_template:
			break
		if "END" in jinja2_template:
			break
		if not sheet.cell(row=i, column=2).value:
			device_names = []
		else:
			device_names = sheet.cell(row=i, column=2).value.splitlines()
		for device in device_names:
			cell_value = {}
			row_data = []
			for j in range(3, nbr_col):
				sheet_row1 = sheet.cell(row=1, column=j).value
				sheet_row = str(sheet.cell(row=i, column=j).value)
				if sheet_row == 'None':
					sheet_row = ''
				cell_value[sheet_row1] = sheet_row
			
			row_data.append(cell_value)
			generate_config(jinja2_template, device, row_data, directory)


# Generate Configuration from "row_data" and append it to output File
def generate_config(jinja2_template: str, device: str, all_rows: list, directory: str):
	jinja2_path = "jinja2_files/" + jinja2_template
	with open(jinja2_path) as t:
		file_jinja2 = Template(t.read())
	replace_file = file_jinja2.render(Variables=all_rows)
	
	file_path = f"{directory}/{device}"
	with open(file_path, "a+", encoding='utf-8') as f:
		f.write(replace_file)
